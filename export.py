import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _style_excel(path, df):
    wb = load_workbook(path)
    ws = wb.active

    headers = df.columns.tolist()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 2):
        for col_num, cell in enumerate(row, 1):
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="DCE6F1")
            if headers[col_num - 1] in ["envoye_ville", "envoye_entreprise", "livre", "facture"]:
                cell.alignment = center_align
            if headers[col_num - 1] in ["montant_engage", "montant_facture", "montant_restant"]:
                cell.number_format = '#,##0.00 €'

    for i, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    wb.save(path)


def _convert_booleans_to_oui_non(df):
    boolean_cols = ["envoye_ville", "envoye_entreprise", "livre", "facture"]
    for col in boolean_cols:
        if col in df.columns:
            df[col] = df[col].map({0: "non", 1: "oui"})
    return df


def export_all_to_excel(path=None):
    try:
        conn = sqlite3.connect("bons.db")
        df = pd.read_sql_query("SELECT * FROM bons_de_commande", conn)
        conn.close()

        df = _convert_booleans_to_oui_non(df)

        if not path:
            path = "bons_export.xlsx"

        df.to_excel(path, index=False)
        _style_excel(path, df)

        print(f"Exporté avec succès dans {path}")

    except Exception as e:
        print("Erreur export Excel :", e)


def export_filtered_to_excel(data, path=None):
    try:
        df = pd.DataFrame(data, columns=[
            "id", "id_bon", "annee_comptable", "numero_bon", "montant_engage", "fournisseur", "imputation",
            "envoye_ville", "envoye_entreprise", "livre", "facture", "numero_facture",
            "montant_facture", "montant_restant", "description", "commentaire"
        ])

        df = _convert_booleans_to_oui_non(df)

        if not path:
            path = "bons_filtres.xlsx"

        df.to_excel(path, index=False)
        _style_excel(path, df)

        print(f"Export filtré effectué dans {path}")
    except Exception as e:
        print("Erreur export filtré :", e)
